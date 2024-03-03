from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook, load_workbook
from tqdm import tqdm
import time


'''
model 3 = https://www.autoscout24.it/lst/tesla/model-3?atype=C&page=1&search_id=15ooxx3ubph&source=listpage_pagination
model y = https://www.autoscout24.it/lst/tesla/model-y?atype=C&page=1&search_id=rt78luelvz&source=listpage_pagination
model s = https://www.autoscout24.it/lst/tesla/model-s?atype=C&page=1&search_id=rhwq5bsgzp&source=listpage_pagination
model x = https://www.autoscout24.it/lst/tesla/model-x?atype=C&page=1&search_id=gyp2wpw5qh&source=listpage_pagination
'''

print ("Caricamento dati...")

def carica_dati (url):

    wb = load_workbook("Cartella.xlsx")
    ws = wb.active


    ws["A1"].value = "Prices"
    ws["B1"].value = "Year"
    ws["C1"].value = "Model"
    ws["D1"].value = "Km"
    ws["E1"].value = "Title"
    ws["F1"].value = "Link"

    html = requests.get(url)
    s = BeautifulSoup(html.content, "html.parser")

    result = s.find(class_ ="ListPage_container__Optya")
    page = int((result.find_all("li", attrs = {"class" : "pagination-item"}))[-1].text.strip())
    idx = 2

    for i in tqdm (range (1, page+1)):

        if i > 10:
            url = url.replace("page=" + url[url.index("page=") + 5] + url[url.index("page=") + 6] , "page=" + str(i))
        else:
            url = url.replace("page=" + url[url.index("page=") + 5], "page=" + str(i))
        html = requests.get(url)
        s = BeautifulSoup(html.content, "html.parser")

        paragrafo = s.find_all("div", attrs = {"class" : "ListItem_wrapper__TxHWu"})
        
        for p in paragrafo:
            if (p.find("p", attrs = {"class" : "Price_price__APlgs PriceAndSeals_current_price__ykUpx"})):
                ws.cell(row=idx, column=1).value = p.find("p", attrs = {"class" : "Price_price__APlgs PriceAndSeals_current_price__ykUpx"}).text.strip()[:-2][1:]
                # print (p.find("p", attrs = {"class" : "Price_price__APlgs PriceAndSeals_current_price__ykUpx"}).text.strip()[:-2])
            else:
                ws.cell(row=idx, column=1).value = p.find("span", attrs = {"class" : "SuperDeal_highlightContainer__R8edU"}).text.strip()[:-3][1:]
                # print (p.find("span", attrs = {"class" : "SuperDeal_highlightContainer__R8edU"}).text.strip()[:-3])

            Km = p.find("span", attrs = {"class" : "VehicleDetailTable_item__4n35N"}).text.strip()
            data = ((p.find_all("span", attrs = {"class" : "VehicleDetailTable_item__4n35N"}))[2]).text.strip()
            CV = (((p.find_all("span", attrs = {"class" : "VehicleDetailTable_item__4n35N"}))[4]).text.strip())[-7:][:6]
            link = p.find("a", attrs = {"class" : "ListItem_title__ndA4s ListItem_title_new_design__QIU2b Link_link__Ajn7I"}).get("href")
            title = p.find("span", attrs = {"class" : "ListItem_version__5EWfi"}).text.strip()


            ws.cell(row=idx, column=2).value = data
            ws.cell(row=idx, column=3).value = CV
            ws.cell(row=idx, column=4).value = Km
            ws.cell(row=idx, column=5).value = title
            ws.cell(row=idx, column=6).value = "https://www.autoscout24.it" + link


            idx += 1
    name = ""
    if "model-3" in url:
        name = "_model_3"
    elif "model-y" in url:
        name = "_model_y"
    elif "model-s" in url:
        name = "_model_s"
    elif "model-x" in url:
        name = "_model_x"
    else:
        print("Errore..")

    t = time.localtime()
    ws["G1"].value = str(time.strftime("%H:%M:%S", t))
    wb.save("output/tesla"+ name +".xlsx") 
       
    
url3 = "https://www.autoscout24.it/lst/tesla/model-3?atype=C&page=1&search_id=15ooxx3ubph&source=listpage_pagination"
urly = "https://www.autoscout24.it/lst/tesla/model-y?atype=C&page=1&search_id=rt78luelvz&source=listpage_pagination"
urls = "https://www.autoscout24.it/lst/tesla/model-s?atype=C&page=1&search_id=rhwq5bsgzp&source=listpage_pagination"
urlx = "https://www.autoscout24.it/lst/tesla/model-x?atype=C&page=1&search_id=gyp2wpw5qh&source=listpage_pagination"



carica_dati(url3)
carica_dati(urly)
carica_dati(urls)
carica_dati(urlx)

print ("Finito!")